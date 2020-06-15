from openpyxl import load_workbook
import os
import xlrd


class FuncOfConvert:
    def __init__(self):
        self.data = {}
        self.house_hold = []  # 法定代表人、经营者家庭资债情况
        self.operation = []  # 资产负债、经营情况
        self.credit = []  # 原有授信担保情况
        self.guarantor = []  # 第二保证人

    def read_data_from_excel(self, excel_file_path):
        wb = load_workbook(excel_file_path, read_only=True)
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]

        self.data['bank_name'] = ws.cell(row=4, column=3).value  # 支行名称 A3's value
        self.data['customer_name'] = ws.cell(row=5, column=2).value  # 客户 B4's value
        self.data['manager_person'] = ws.cell(row=4, column=13).value  # 归管客户经理 A3's value

        for row_number in range(1, ws.max_row + 1):
            value = str(ws.cell(row_number, 1).value)
            if '企业情况说明' in value and '企业基本情况' in value:
                self.data['customer_basic_info_1'] = ws.cell(row=row_number + 1, column=1).value  # 借款人基本情况1
                continue
            if '企业征信及外部查询负面信息' in value:
                self.data['customer_basic_info_2'] = ws.cell(row=row_number + 1, column=1).value  # 借款人基本情况2
                continue
            if '关联企业情况（含企业股东、控股子公司及其他实质关联企业）' in value:
                self.data['associate_enterprise_info'] = ws.cell(row=row_number + 1, column=1).value  # 关联企业情况
                continue
            if '关联并表情况' in value:
                self.data['associate_merge_table'] = ws.cell(row=row_number + 1, column=1).value  # 关联并表
                self.house_hold.append(row_number + 2)
                continue
            if '法定代表人（实际经营者）及配偶相关说明' in value:
                self.data['enterprise_operator_info_1'] = ws.cell(row=row_number + 1, column=1).value  # 企业经营者相关情况1
                continue
            if '法定代表人（实际经营者）及配偶征信及外部查询负面信息' in value:
                self.data['enterprise_operator_info_2'] = ws.cell(row=row_number + 1, column=1).value  # 企业经营者相关情况2
                continue
            if '企业财务情况' in value:
                self.data['enterprise_finance_condition_1'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况1
                continue
            if '相关情况说明' in value:
                self.data['enterprise_finance_condition_2'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况2
                self.operation.append(row_number - 1)
                continue
            if '具体描述人品、产品、抵押品；电表、水表、纳税报表情况' in value:
                self.data['enterprise_finance_condition_3'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况3
                continue
            if '抵押物及保证人情况介绍' in value:
                self.data['warrantor_and_guaranty_1'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                self.credit.append(row_number - 1)
                continue
            if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                self.data['warrantor_and_guaranty_2'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                continue
            if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                self.data['warrantor_and_guaranty_2'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                continue
            if '申报授信理由：' in value:
                self.data['declaration_reason_and_purpose_1'] = ws.cell(row=row_number + 1, column=1).value
                continue
                # 支行申报理由及用途1
            if '本次贷款用途及第一还款来源分析：' in value:
                self.data['declaration_reason_and_purpose_2'] = ws.cell(row=row_number + 1, column=1).value
                continue
                # 支行申报理由及用途2
            if '客户信用记录（若无即填' in value:
                self.house_hold.append(row_number - 1)
                continue
            if '资产负债、经营情况（保留到个位数）' in value:
                self.operation.append(row_number)
                continue
            if '原有授信担保情况' == value:
                self.credit.append(row_number)
                continue
            if '第二保证人落实情况' in value:
                self.guarantor.append(row_number + 1)
                continue
            if '上次授信批复要求及贷后管理情况（摘抄前次批复详细内容' in value:
                self.guarantor.append(row_number - 1)
                continue
        for key in self.data.keys():
            print(key, self.data[key])
        print(self.house_hold)
        print(self.operation)
        print(self.guarantor)
        print(self.credit)


if __name__ == "__main__":
    func = FuncOfConvert()
    func.read_data_from_excel('/Users/fangyong/Desktop/2020模板审核调查报告——威远机电.xlsx')
