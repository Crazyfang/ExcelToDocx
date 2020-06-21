from win32com import client
from openpyxl import load_workbook
import os
import xlrd
import logging
import time
import pythoncom
import sys
sys.coinit_flags = 0


def generate_logging():
    """
    return:
        The logger output the log message
    """
    # First, generate a logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Second, generate a log handler to write the log file
    rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
    log_path = os.getcwd() + '/Logs/'
    if os.path.exists(log_path):
        pass
    else:
        os.mkdir(log_path)

    log_name = log_path + rq + '.log'
    logfile = log_name
    fh = logging.FileHandler(logfile, mode='w')
    fh.setLevel(logging.DEBUG)

    # Third, define the output format of handler
    formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
    fh.setFormatter(formatter)

    # Four, add the handler into logger
    logger.addHandler(fh)

    return logger


class FuncOfConvert:
    def __init__(self):
        self.data = {}
        self.house_hold = []  # 法定代表人、经营者家庭资债情况
        self.operation = []  # 资产负债、经营情况
        self.credit = []  # 原有授信担保情况
        self.guarantor = []  # 第二保证人
        self.logger = generate_logging()
        print('参数读取')
        pythoncom.CoInitialize()
        self.excel = client.Dispatch('Excel.Application', clsctx=pythoncom.CLSCTX_LOCAL_SERVER)
        self.word = client.Dispatch('Word.Application', clsctx=pythoncom.CLSCTX_LOCAL_SERVER)
        print('参数读取成功')

    def get_file_list(self, file_path):
        file_list = []
        for file_name in os.listdir(file_path):
            extension = os.path.splitext(file_name)[-1][1:]
            if extension == 'xlsx' or extension == 'xls':
                file_list.append(os.path.join(file_path, file_name))

        return file_list

    def read_data_from_excel(self, excel_file_path):
        try:
            wb = load_workbook(excel_file_path, read_only=True)
            sheet_names = wb.sheetnames
            ws = wb[sheet_names[0]]

            self.data = {}
            self.house_hold = []  # 法定代表人、经营者家庭资债情况
            self.operation = []  # 资产负债、经营情况
            self.credit = []  # 原有授信担保情况
            self.guarantor = []  # 第二保证人


            self.data['bank_name'] = ws.cell(row=4, column=3).value  # 支行名称 A3's value
            self.data['customer_name'] = ws.cell(row=5, column=2).value  # 客户 B4's value
            self.data['manager_person'] = ws.cell(row=4, column=13).value  # 归管客户经理 A3's value

            for row_number in range(1, ws.max_row + 1):
                value = str(ws.cell(row_number, 1).value)
                if '企业情况说明' in value and '企业基本情况' in value:
                    self.data['customer_basic_info_1'] = ws.cell(row=row_number + 1, column=1).value  # 借款人基本情况1
                    continue
                if '企业征信及外部查询负面信息' in value:
                    self.data['customer_basic_info_2'] = ws.cell(row=row_number + 1, column=1).value  # 借款人基本情况2
                    continue
                if '关联企业情况（含企业股东、控股子公司及其他实质关联企业）' in value:
                    self.data['associate_enterprise_info'] = ws.cell(row=row_number + 1, column=1).value  # 关联企业情况
                    continue
                if '关联并表情况' in value:
                    self.data['associate_merge_table'] = ws.cell(row=row_number + 1, column=1).value  # 关联并表
                    self.house_hold.append(row_number + 2)
                    continue
                if '法定代表人（实际经营者）及配偶相关说明' in value:
                    self.data['enterprise_operator_info_1'] = ws.cell(row=row_number + 1, column=1).value  # 企业经营者相关情况1
                    continue
                if '法定代表人（实际经营者）及配偶征信及外部查询负面信息' in value:
                    self.data['enterprise_operator_info_2'] = ws.cell(row=row_number + 1, column=1).value  # 企业经营者相关情况2
                    continue
                if '企业财务情况' in value:
                    self.data['enterprise_finance_condition_1'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况1
                    continue
                if '相关情况说明' in value:
                    self.data['enterprise_finance_condition_2'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况2
                    self.operation.append(row_number - 1)
                    continue
                if '具体描述人品、产品、抵押品；电表、水表、纳税报表情况' in value:
                    self.data['enterprise_finance_condition_3'] = ws.cell(row=row_number + 1, column=1).value  # 企业财务状况3
                    continue
                if '抵押物及保证人情况介绍' in value:
                    self.data['warrantor_and_guaranty_1'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                    self.credit.append(row_number - 1)
                    continue
                if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                    self.data['warrantor_and_guaranty_2'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                    continue
                if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                    self.data['warrantor_and_guaranty_2'] = ws.cell(row=row_number + 1, column=1).value  # 保证人及抵押物情况
                    continue
                if '申报授信理由：' in value:
                    self.data['declaration_reason_and_purpose_1'] = ws.cell(row=row_number + 1, column=1).value
                    continue
                    # 支行申报理由及用途1
                if '本次贷款用途及第一还款来源分析：' in value:
                    self.data['declaration_reason_and_purpose_2'] = ws.cell(row=row_number + 1, column=1).value
                    continue
                    # 支行申报理由及用途2
                if '客户信用记录（若无即填' in value:
                    self.house_hold.append(row_number - 1)
                    continue
                if '资产负债、经营情况（保留到个位数）' in value:
                    self.operation.append(row_number)
                    continue
                if '原有授信担保情况' == value:
                    self.credit.append(row_number)
                    continue
                if '第二保证人落实情况' in value:
                    self.guarantor.append(row_number + 1)
                    continue
                if '上次授信批复要求及贷后管理情况（摘抄前次批复详细内容' in value:
                    self.guarantor.append(row_number - 1)
                    continue
        except Exception as e:
            self.logger.error(str(e))

    def read_data_from_xls(self, excel_file_path):
        try:
            excel = xlrd.open_workbook(excel_file_path)
            ws = excel.sheet_by_index(0)


            self.data = {}
            self.house_hold = []  # 法定代表人、经营者家庭资债情况
            self.operation = []  # 资产负债、经营情况
            self.credit = []  # 原有授信担保情况
            self.guarantor = []  # 第二保证人

            self.data['bank_name'] = ws.cell(rowx=3, colx=2).value  # 支行名称 A3's value
            self.data['customer_name'] = ws.cell(rowx=4, colx=1).value  # 客户 B4's value
            self.data['manager_person'] = ws.cell(rowx=3, colx=12).value  # 归管客户经理 A3's value

            print(ws.nrows)

            for row_number in range(0, ws.nrows):
                value = str(ws.cell(row_number, 0).value)
                if '企业情况说明' in value and '企业基本情况' in value:
                    self.data['customer_basic_info_1'] = ws.cell(rowx=row_number + 1, colx=0).value  # 借款人基本情况1
                    continue
                if '企业征信及外部查询负面信息' in value:
                    self.data['customer_basic_info_2'] = ws.cell(rowx=row_number + 1, colx=0).value  # 借款人基本情况2
                    continue
                if '关联企业情况（含企业股东、控股子公司及其他实质关联企业）' in value:
                    self.data['associate_enterprise_info'] = ws.cell(rowx=row_number + 1, colx=0).value  # 关联企业情况
                    continue
                if '关联并表情况' in value:
                    self.data['associate_merge_table'] = ws.cell(rowx=row_number + 1, colx=0).value  # 关联并表
                    self.house_hold.append(row_number + 3)
                    continue
                if '法定代表人（实际经营者）及配偶相关说明' in value:
                    self.data['enterprise_operator_info_1'] = ws.cell(rowx=row_number + 1, colx=0).value  # 企业经营者相关情况1
                    continue
                if '法定代表人（实际经营者）及配偶征信及外部查询负面信息' in value:
                    self.data['enterprise_operator_info_2'] = ws.cell(rowx=row_number + 1, colx=0).value  # 企业经营者相关情况2
                    continue
                if '企业财务情况' in value:
                    self.data['enterprise_finance_condition_1'] = ws.cell(rowx=row_number + 1, colx=0).value  # 企业财务状况1
                    continue
                if '相关情况说明' in value:
                    self.data['enterprise_finance_condition_2'] = ws.cell(rowx=row_number + 1, colx=0).value  # 企业财务状况2
                    self.operation.append(row_number)
                    continue
                if '具体描述人品、产品、抵押品；电表、水表、纳税报表情况' in value:
                    self.data['enterprise_finance_condition_3'] = ws.cell(rowx=row_number + 1, colx=0).value  # 企业财务状况3
                    continue
                if '抵押物及保证人情况介绍' in value:
                    self.data['warrantor_and_guaranty_1'] = ws.cell(rowx=row_number + 1, colx=0).value  # 保证人及抵押物情况
                    self.credit.append(row_number)
                    continue
                if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                    self.data['warrantor_and_guaranty_2'] = ws.cell(rowx=row_number + 1, colx=0).value  # 保证人及抵押物情况
                    continue
                if '保证人情况须阐述保证人基本情况、净资产、经营收入及担保能力情况' in value:
                    self.data['warrantor_and_guaranty_2'] = ws.cell(rowx=row_number + 1, colx=0).value  # 保证人及抵押物情况
                    continue
                if '申报授信理由：' in value:
                    self.data['declaration_reason_and_purpose_1'] = ws.cell(rowx=row_number + 1, colx=0).value
                    continue
                    # 支行申报理由及用途1
                if '本次贷款用途及第一还款来源分析：' in value:
                    self.data['declaration_reason_and_purpose_2'] = ws.cell(rowx=row_number + 1, colx=0).value
                    continue
                    # 支行申报理由及用途2
                if '客户信用记录（若无即填' in value:
                    self.house_hold.append(row_number)
                    continue
                if '资产负债、经营情况（保留到个位数）' in value:
                    self.operation.append(row_number + 1)
                    continue
                if '原有授信担保情况' == value:
                    self.credit.append(row_number + 1)
                    continue
                if '第二保证人落实情况' in value:
                    self.guarantor.append(row_number + 2)
                    continue
                if '上次授信批复要求及贷后管理情况（摘抄前次批复详细内容' in value:
                    self.guarantor.append(row_number)
                    continue
        except Exception as e:
            self.logger.error(str(e))

    def win32test(self, excel_file_path):
        try:
            doc_file_path = os.path.splitext(excel_file_path)[0] + '.docx'
            doc = self.word.Documents.Add()
            book = self.excel.Workbooks.Open(excel_file_path)

            sheet = book.Worksheets(1)
            # sheet.Range('A36:AE44').Copy()
            sheet.Range('A{0}:AE{1}'.format(*self.house_hold)).Copy()

            # myRange = doc.Range()
            # myRange = doc.Selection
            self.word.Selection.InsertAfter(self.data['bank_name'] + '：' + self.data['customer_name'] + '  归管客户经理：' + self.data[
                'manager_person'] + '\n')
            self.word.Selection.InsertAfter('(一)借款人基本情况\n')
            self.word.Selection.InsertAfter(self.data['customer_basic_info_1'])
            self.word.Selection.InsertAfter(self.data['customer_basic_info_2'])
            self.word.Selection.InsertAfter('关联企业情况：' + self.data['associate_enterprise_info'])
            self.word.Selection.InsertAfter('关联并表：' + self.data['associate_merge_table'])
            self.word.Selection.InsertAfter('(二)企业经营者相关情况\n')
            self.word.Selection.InsertAfter(self.data['enterprise_operator_info_1'])
            self.word.Selection.InsertAfter(self.data['enterprise_operator_info_2'])

            # wdRange = doc.Content
            # wdRange.Collapse(0)
            self.word.Selection.MoveRight()
            self.word.Selection.PasteExcelTable(False, False, False)
            # wdRange.PasteExcelTable(False, False, False)

            self.word.Selection.InsertAfter('(三)企业财务状况\n')
            self.word.Selection.InsertAfter(self.data['enterprise_finance_condition_1'])

            # sheet.Range('A108:AE157').Copy()
            sheet.Range('A{0}:AE{1}'.format(*self.operation)).Copy()
            self.word.Selection.MoveRight()
            self.word.Selection.PasteExcelTable(False, False, False)
            # wdRange.PasteExcelTable(False, False, False)

            self.word.Selection.InsertAfter(self.data['enterprise_finance_condition_2'])
            self.word.Selection.InsertAfter(self.data['enterprise_finance_condition_3'])

            self.word.Selection.InsertAfter('(四)存量授信及申报授信情况\n')

            # sheet.Range('A62:AE86').Copy()
            sheet.Range('A{0}:AE{1}'.format(*self.credit)).Copy()
            self.word.Selection.MoveRight()
            self.word.Selection.PasteExcelTable(False, False, False)
            # wdRange.PasteExcelTable(False, False, False)

            self.word.Selection.InsertAfter('保证人及抵押物情况介绍\n')
            self.word.Selection.InsertAfter(self.data['warrantor_and_guaranty_1'])
            self.word.Selection.InsertAfter(self.data['warrantor_and_guaranty_2'])

            self.word.Selection.InsertAfter('第二保证人落实情况：\n')
            # sheet.Range('A161:AE166').Copy()
            sheet.Range('A{0}:AE{1}'.format(*self.guarantor)).Copy()
            self.word.Selection.MoveRight()
            self.word.Selection.PasteExcelTable(False, False, False)
            # wdRange.PasteExcelTable(False, False, False)

            self.word.Selection.InsertAfter('(五)支行申报理由及用途\n')
            self.word.Selection.InsertAfter(self.data['declaration_reason_and_purpose_1'])
            self.word.Selection.InsertAfter(self.data['declaration_reason_and_purpose_2'])

            self.word.Selection.InsertAfter('授信部意见：\n')
            self.word.Selection.InsertAfter('风险提示：\n')
            self.word.Selection.InsertAfter('(六)授信审批委员会集体审议结论\n')

            doc.SaveAs(doc_file_path)
            doc.Close()

            book.Application.CutCopyMode = False
            book.Close()

            self.data.clear()

            print('转移完成!')
        except Exception as e:
            self.logger.error(str(e))


if __name__ == '__main__':
    excel_path = 'C:\\Users\\Administrator\\Desktop\\2020模板审核调查报告——威远机电.xls'
    word_path = 'C:\\Users\\Administrator\\Desktop\\123.docx'

    cla = FuncOfConvert()
    cla.read_data_from_xls(excel_path)
    cla.win32test(excel_path)
    # cla.test()
    # cla.win32test()
